
'''
Handles the opening, creation, reading and writing of excel files. As it 
copies data from an excel file it does some basic sanitation specific to 
this application. 
'''

from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.styles import Color
from openpyxl.styles import Font
from openpyxl.styles import PatternFill

from pathlib import Path


class ReadWriteExcel:

    def __init__(self, path=''):
        self.path = path
        # if workbook exist load it if not create it
        file = Path(self.path)
        if file.is_file():
            self.wb = load_workbook(self.path)
        else: 
            self.wb = Workbook()

        self.ws = self.wb.active
        self.max_row = self.ws.max_row
        self.max_col = self.ws.max_column

    def copy_row(self, row_to_copy):
        # copy a given row into a list() each cell value is a list value
        row_list = list()
        for col in range(1, (self.max_col + 1)):
            row_list.append(self.ws.cell(row=row_to_copy, column=col).value)
        row_list = self.clean_list(row_list)
        return row_list

    def copy_col(self, col_to_copy, start_row):
        # copy a given column into a list() each cell value is a list value
        col_list = list()
        for row in range(start_row, (self.max_row + 1)):
            col_list.append(self.ws.cell(row=row, column=col_to_copy).value)
            col_list = self.clean_list(col_list)
        return col_list

    def find_empty_col(self, row):
        # find an empty column in a given row
        col = 1
        while self.ws.cell(row=row, column=col).value is not None:
            col += 1
        return col

    def find_empty_row(self, col):
        # find a empty row in a given column
        row = 1
        while self.ws.cell(row=row, column=col).value is not None:
            row += 1
        return row

    def clean_list(self, dirty_list):
        # sanitize a list so that all items are lowercase strings
        clean_list = [self.clean_str(item) for item in dirty_list]
        return clean_list

    def clean_str(self, dirty_str):
        # sanitize a value so that it is a lowercase string
        if dirty_str is None:
            clean_string = ''
        else:
            dirty_str = str(dirty_str)
            clean_string = dirty_str.strip().lower()
        return clean_string

    def create_worksheets(self, ws_list):
        # create worksheets in the workbook
        # renames the active worksheet to ws_list[0]
        # tried to use ws_list.pop(0) but sheet title would not change
        for index, sheet in enumerate(ws_list):
            if index == 0:
                self.ws.title = sheet
            else:
                self.wb.create_sheet(sheet)

    def write_row(self, row_list, bold=False, row=None):
        # write a row with each list() item as a cell value
        # The optional row value makes it possible to overwrite a row. 
        if not row:
        	row = self.find_empty_row(1)
        col = 1
        for val in row_list:
            self.ws.cell(column=col, row=row, value=val.title()).font = Font(bold=bold)
            col += 1
        

    def highlight_cell(self, row, col):
        # highlight a given cell yellow
        yellowFill = PatternFill(start_color='FFFF4C',
                                 end_color='FFFF4C', fill_type='solid')
        self.ws.cell(row=row, column=col).fill = yellowFill

    def save_workbook(self):
        # if only reading workbook this is not necessary
        self.wb.save(self.path)
